"""
build_artifacts.py
------------------
Reads data_dictionary.md and produces:
  - data_dictionary.xlsx  (one sheet per table + Table Index)
  - mermaid_flow.png      (ER diagram rendered via mermaid.ink)

Usage:
    python3 build_artifacts.py
"""

import re
import base64
import pathlib
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image
import io

SRC = pathlib.Path(__file__).parent / "data_dictionary.md"
OUT_XLSX = pathlib.Path(__file__).parent / "data_dictionary.xlsx"
OUT_PNG = pathlib.Path(__file__).parent / "mermaid_flow.png"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


# ── Markdown parser ───────────────────────────────────────────────────────────

def parse_md(path: pathlib.Path):
    text = path.read_text(encoding="utf-8")

    # Extract mermaid block
    mermaid_match = re.search(r"```mermaid\n(.*?)```", text, re.DOTALL)
    mermaid_code = mermaid_match.group(1).strip() if mermaid_match else ""

    # Extract table index (first markdown table after "## Table Index")
    index_rows = []
    in_index = False
    for line in text.splitlines():
        if "## Table Index" in line:
            in_index = True
            continue
        if in_index:
            if line.startswith("##"):
                break
            if line.startswith("|") and not re.match(r"\|[-| ]+\|", line):
                cells = [c.strip() for c in line.strip("|").split("|")]
                if cells[0] != "#":          # skip header row
                    index_rows.append(cells)

    # Extract per-table sections
    # Each section starts with "## N. `table_name`"
    section_pattern = re.compile(
        r"## \d+\. `(\w+)`.*?\n"           # heading → table name
        r"\*\*EN:\*\* (.*?)\n"
        r"\*\*TH:\*\* (.*?)\n"
        r"\nSource: `(.*?)`",
        re.DOTALL,
    )

    # markdown table rows inside each section
    table_blocks = re.split(r"\n---\n", text)

    tables = []
    for block in table_blocks:
        m = re.search(
            r"## \d+\. `(\w+)`[^\n]*\n\n"
            r"\*\*EN:\*\* (.*?)\n"
            r"\*\*TH:\*\* (.*?)\n"
            r"\nSource: `(.*?)`",
            block,
        )
        if not m:
            continue
        name, en, th, source = m.group(1), m.group(2), m.group(3), m.group(4)

        # parse markdown table rows (skip header + separator)
        rows = []
        in_table = False
        for line in block.splitlines():
            if line.startswith("|"):
                cells = [c.strip() for c in line.strip("|").split("|")]
                if re.match(r"^[-:]+$", cells[0]):  # separator row
                    in_table = True
                    continue
                if in_table:
                    rows.append(cells)
            elif in_table and not line.startswith("|"):
                in_table = False

        tables.append({"name": name, "en": en, "th": th, "source": source, "rows": rows})

    return index_rows, tables, mermaid_code


# ── Excel builder ─────────────────────────────────────────────────────────────

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)


def build_xlsx(index_rows, tables, out_path):
    wb = Workbook()
    wb.remove(wb.active)

    # Table Index sheet
    ws = wb.create_sheet("Table Index")
    ws.append(["#", "Original Sheet Name", "Suggested Table Name", "Domain"])
    style_header(ws)
    for i, row in enumerate(index_rows, 1):
        ws.append([i] + row[1:])   # replace "#" cell with sequential int
    auto_width(ws)

    COL_HEADERS = ["Field (Original)", "Suggested Name", "EN Description",
                   "TH Description", "Type", "Key"]

    for tbl in tables:
        ws = wb.create_sheet(tbl["name"])
        ws.append(["EN:", tbl["en"]]);     ws["A1"].font = Font(bold=True)
        ws.append(["TH:", tbl["th"]]);     ws["A2"].font = Font(bold=True)
        ws.append(["Source:", tbl["source"]]); ws["A3"].font = Font(bold=True)
        ws.append([])
        ws.append(COL_HEADERS)
        style_header(ws, row=5)
        for row in tbl["rows"]:
            ws.append(row)
        auto_width(ws)

    wb.save(out_path)
    print(f"[xlsx] saved → {out_path}")


# ── Mermaid PNG via mermaid.ink ───────────────────────────────────────────────

def build_png(mermaid_code: str, out_path: pathlib.Path):
    render_code = """%%{init: {
      'theme': 'base',
      'themeVariables': {
        'background': '#FFFFFF',
        'primaryColor': '#F8F8F8',
        'primaryTextColor': '#111111',
        'primaryBorderColor': '#222222',
        'lineColor': '#222222',
        'secondaryColor': '#F3F3F3',
        'tertiaryColor': '#FFFFFF',
        'clusterBkg': '#FFFFFF',
        'clusterBorder': '#444444',
        'fontFamily': 'Arial'
      }
    }}%%
""" + mermaid_code

    encoded = base64.urlsafe_b64encode(render_code.encode("utf-8")).decode("ascii")
    url = f"https://mermaid.ink/img/{encoded}?type=png&width=2000&theme=base"

    resp = requests.get(url, timeout=30)
    resp.raise_for_status()

    img = Image.open(io.BytesIO(resp.content)).convert("RGBA")

    # Force a white background for any transparent pixels
    white_bg = Image.new("RGBA", img.size, "white")
    merged = Image.alpha_composite(white_bg, img).convert("RGB")

    merged.save(out_path, "PNG")
    print(f"[png]  saved → {out_path}  ({merged.size[0]}×{merged.size[1]})")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"Reading {SRC} ...")
    index_rows, tables, mermaid_code = parse_md(SRC)
    print(f"  found {len(tables)} tables, mermaid block: {len(mermaid_code)} chars")

    build_xlsx(index_rows, tables, OUT_XLSX)
    build_png(mermaid_code, OUT_PNG)

    print("Done.")
